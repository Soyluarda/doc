from flask import Blueprint, render_template
from flask_login import login_required, current_user
from flask import request, redirect, session, jsonify, url_for
from openpyxl import load_workbook
from .models import Doctors


main = Blueprint('main', __name__)

@main.route('/')
def index():
    return render_template('index.html')

@main.route('/upload-file')
@login_required
def upload_file():
    return render_template('upload_file.html')

@main.route('/check', methods=['POST'])
def check():
    wb = load_workbook(filename=request.files['file'])
    ws = wb['Sheet1']

    list_with_values = []
    for cell in ws[1]:
        list_with_values.append(cell.value)

    necessary_columns = set(['doctor_name', 'doctor_surname'])
    if necessary_columns.issubset(list_with_values):
        return jsonify({'status': 'success'})

    return jsonify({'status': 'error'})


@main.route('/process', methods=['POST'])
def process():
    wb = load_workbook(filename=request.files['file'])
    ws = wb['Sheet1']
    doctors = Doctors.query.with_entities(Doctors.name, Doctors.surname)
    message = 'Matched'

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row not in doctors:
            message = 'Not Matched'

    return jsonify({'status': message})